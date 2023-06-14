import imp
from msilib.schema import Class
from pyexpat import features
from statistics import mode
import torch.nn as nn
import warnings
import tqdm
import torch.optim as optim
import torchvision.transforms as transforms
import torch.nn.functional as F
import os
import openpyxl
import glob
import pprint as pp
import cv2
import numpy as np
import csv
import torch
import shutil
import matplotlib.pyplot as plt
from torch.utils.data import Dataset
from torchvision import datasets, transforms
from torch.utils.data import Dataset, DataLoader
from torch.utils.data.dataset import Subset
from sklearn.model_selection import StratifiedKFold
from torchvision import models
from gradcam.utils import visualize_cam
from PIL import Image
from torchcam.methods import GradCAMpp
from itertools import chain
from torchcam.utils import overlay_mask
from torchvision import models
from torchvision.io.image import read_image
from torchvision.transforms.functional import normalize, resize, to_pil_image

# モデルのパスとデータセットのパス、出力ファイル名を指定
modelPath = r"F:\研究関連\TwoStreamCNN_2nd-Season\Models\MobileNet_C-CoOcc_AROB-Journal\MobileNet_C-CoOcc_AROB-Journal_max-Acc.pth"
dataset = r"F:\AROB_Journal\20BN_Jester-2000Grad\C-CoOcc"
modelName = r"C-CoOcc_AROB-Journal_max-Acc"

# 画像の前処理を定義
transform = transforms.Compose([
    transforms.Resize((224, 224), interpolation=transforms.InterpolationMode.BICUBIC),
    transforms.ToTensor(),
    transforms.Normalize(
        mean=[0.5, 0.5, 0.5],
        std=[0.5, 0.5, 0.5])
])

# MobileNetモデルの定義
class MobileNet():
    def __init__(self):
        self.model = models.mobilenet_v2(pretrained=True)
        self.model.classifier[1] = nn.Linear(1280, 6)
        self.transGrad(True)

    def transGrad(self, Boo):
        for p in self.model.features.parameters():
            p.requires_grad = Boo

# Grad-CAMを実行するクラス
class Cam():
    def __init__(self, loadmodel, DataPath):
        self.model = loadmodel
        self.device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
        self.cam_extractor = GradCAMpp(self.model.eval())
        self.ClassList = os.listdir(DataPath)
    
    def GradCam(self, image, label, org_image):
        output = self.model((image.unsqueeze(0)).to(self.device))
        pred = (F.softmax((self.model((image.unsqueeze(dim=0)).to(self.device)))))
        pred_idx = pred.max(1)[1]
        cam = self.cam_extractor(output.squeeze(0).argmax().item(), output)
        result = overlay_mask(org_image, to_pil_image(cam[0], mode="F"), alpha=0.5)
        heatmap = overlay_mask(Image.new("RGB", (224,224), color=(0,0,0)),to_pil_image(cam[0], mode="F"),alpha=0.001)
        pred_value = pred.data[0].tolist()
        return result, pred_idx, pred_value, heatmap
    
    def __call__(self, image, label):
        result_img, pred_idx, pred_value = self.GradCam(image, label)

# Excelファイルに結果を書き込むクラス
class ExcelWriter:
    def __init__(self, classlist):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.worksheets[0]
        self.ws.title = modelName
        self.classList = classlist
        self.columnName()

    def columnName(self):
        names = ["画像名", "ラベル", "推論結果"] + self.classList + ["オリジナル画像", "GradCam"]
        [self.ws.cell(row=1, column=x+1, value=name) for x, name in enumerate(names)]
    
    def writeWorkbook(self, index, datas, images):
        [self.ws.cell(row=index+2, column=x+1, value=data) for x, data in enumerate(datas)]
        self.ws.row_dimensions[index+2].height = 224*0.75
        for i, image in enumerate(images):
            column_leter = openpyxl.utils.cell.get_column_letter(self.ws.cell(row=index+2, column=(len(datas)+i+1)).column)
            self.ws.column_dimensions[column_leter].width = 224*0.13
            cell_address = self.ws.cell(row=index+2, column=len(datas)+1+i).coordinate
            img = openpyxl.drawing.image.Image(image)
            img.anchor = cell_address
            self.ws.add_image(img)
    
    def _saveWorkbook(self, path):
        self.wb.save(path)

if __name__ == '__main__':
    Mobilenet = MobileNet()
    model = Mobilenet.model.cuda()
    checkpoint = torch.load(modelPath)
    model.load_state_dict(checkpoint['model_state_dict'])
    ClassList = os.listdir(dataset)
    Gradcam = Cam(loadmodel=model, DataPath=dataset)
    Data = datasets.ImageFolder(root=dataset, transform=transform)
    ImagePaths = list(chain.from_iterable([glob.glob('{}/*.jpg'.format(d)) for d in glob.glob('{}/*'.format(dataset))]))
    os.mkdir(modelName)
    os.mkdir(str(modelName)+"/GradCam_Result")
    os.mkdir(str(modelName)+"/Original_Image")
    shutil.copyfile("template.xlsx",str(modelName)+"/"+str(modelName)+".xlsx")
    ImageNo = 0
    excel = ExcelWriter(ClassList)
    for images, labels in (Data):
        org_image = Image.open(ImagePaths[ImageNo]).resize((224, 224),Image.BICUBIC)
        Gradcam_img, pred_idx, pred_value, heatmap = Gradcam.GradCam(image=images, label=labels, org_image = org_image)
        gradImagePath = str(modelName)+"/GradCam_Result/"+str(ClassList[labels])+"_"+str(ImageNo)+".jpg"
        orgImagePath = str(modelName)+"/Original_Image/"+str(ClassList[labels])+"_"+str(ImageNo)+".jpg"
        Gradcam_img.save(gradImagePath)
        org_image.save(orgImagePath)
        inputData = [os.path.basename(ImagePaths[ImageNo]), os.path.basename(os.path.dirname(ImagePaths[ImageNo])), ClassList[pred_idx]] + [i for i in pred_value]
        inputImage = [orgImagePath, gradImagePath]
        excel.writeWorkbook(ImageNo, inputData, inputImage)
        ImageNo += 1
    excel._saveWorkbook(str(modelName)+"/"+str(modelName)+".xlsx")
