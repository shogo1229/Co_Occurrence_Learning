#--------------------------------------------------------------------------------#
#これは領家産のコードであり無能な由田君はこのような美しいコードを生産することはできません#
#--------------------------------------------------------------------------------#
import urllib
import pickle
import cv2
import numpy as np
from PIL import Image
import glob
import os
import math
import torch
import torch.nn as nn
import torch.nn.functional as F
from torchvision import transforms
from torchvision import models
from itertools import chain
import matplotlib.pyplot as plt
import seaborn as sns
import imghdr
import openpyxl
from PIL import JpegImagePlugin
JpegImagePlugin._getmp = lambda x: None

Height = 224 # 画像の高さ(このサイズにリサイズされる)
Width = 224 # 画像の幅(このサイズにリサイズされる)
# imgPaths = r"D:\NaoChin\windowsold\ducuments\dataset\VGG\fishery\test" # 推論テスト用
# imgPaths = r"D:\NaoChin\windowsold\ducuments\code\Naochin_BingObjectness\NaochinObjectness\Objectness\ResVOC\014171\Obj"
imgPaths = r"E:\Research\DataSet\UCF11\test\Synthetic_100_resize"
# modelPath = r"model.pth" # モデルを保存するパス
modelPath = r"E:\Research\TwoStreamCNN_2nd-Season\max_Val_acc_VGG16_UCF11_Synthec_part2.pth" # モデルを保存するパス
transform = transforms.Compose([
			transforms.Resize((Height, Width), interpolation=Image.BICUBIC),
			transforms.ToTensor()
			])
# classLabel = ["other", "trawl_and_purse_seiner"]
classList = os.listdir(imgPaths)
resultFile = "result.xlsx"
# resultFile = "result/result.xlsx"
# resultFile = r"D:\NaoChin\windowsold\ducuments\code\Naochin_BingObjectness\NaochinObjectness\Objectness\ResVOC\014171\result.xlsx"


# VGG16で転移学習
class VGG():
	def __init__(self):
		self.model = models.vgg16_bn(pretrained=True)
		self.transformClassifier()
		self.transGrad(True)
		
	# 全結合層の変形(最終出力層を学習クラス数に変形する必要があるため)
	def transformClassifier(self):
		self.model.classifier = nn.Sequential(
			nn.Linear(512 * 7 * 7, 4096),
			nn.ReLU(True),
			nn.Dropout(),
			nn.Linear(4096, 4096),
			nn.ReLU(True),
			nn.Dropout(),
			nn.Linear(4096, 11)
			)
	# 微分するかしないかみたいなやつ
	def transGrad(self, Boo):
		for p in self.model.features.parameters():
			p.requires_grad = Boo

class GradCam:
	def __init__(self, model):
		self.model = model
		self.model.eval()
		self.transform = transform
		self.feature = None
		self.gradient = None
		self.testImage = None
	
	def save_gradient(self, grad):
		# print(grad)
		self.gradient = grad

	def __call__(self, x):
		image_size = (x.size(-1), x.size(-2))
		feature_maps = []

		for i in range(x.size(0)):
			img = x[i].data.cpu().numpy()
			img = img - np.min(img)
			if np.max(img) != 0:
				img = img / np.max(img)

			feature = x[i].unsqueeze(0)

			# print(type(feature))

			for name, module in self.model.named_children():
				#print(name, module)
				if name == 'classifier':
					feature = feature.view(feature.size(0), -1)
				feature = module(feature)
				#print(feature)
				if name == 'features':
					feature.register_hook(self.save_gradient)
					self.feature = feature

				classes = F.sigmoid(feature)
			one_hot, _ = classes.max(dim=-1) # こっちだと識別したクラスの特徴量を出力
			# one_hot = classes[0, math.ceil((now+1)/10)-1] # こっちだと正解のクラスの特徴量を出力
			self.model.zero_grad()
			one_hot.backward()
			weight = self.gradient.mean(dim=-1, keepdim=True).mean(dim=-2, keepdim=True)

			mask = F.relu((weight * self.feature).sum(dim=1)).squeeze(0)
			mask = cv2.resize(mask.data.cpu().numpy(), image_size)
			mask = mask - np.min(mask)

			if np.max(mask) != 0:
				mask = mask / np.max(mask)

			feature_map = np.float32(cv2.applyColorMap(np.uint8(255 * mask), cv2.COLORMAP_JET))
			cam = feature_map + np.float32((np.uint8(img.transpose((1, 2, 0)) * 255)))
			cam = cam - np.min(cam)

			if np.max(cam) != 0:
				cam = cam / np.max(cam)

			feature_maps.append(transforms.ToTensor()(cv2.cvtColor(np.uint8(255 * cam), cv2.COLOR_BGR2RGB)))

		feature_maps = torch.stack(feature_maps)

		return feature_maps

class ExcelWriter:
	def __init__(self):
		# 実行結果の保存の準備
		self.wb = openpyxl.Workbook()
		self.ws = self.wb.worksheets[0]
		self.ws.title = "result+grad_cam"
		self.columnName()

	def columnName(self):
		names = ["画像名", "ラベル", "推論結果"] + classList + ["オリジナル画像", "GradCam"]
		[self.ws.cell(row=1, column=x+1, value=name)for x, name in enumerate(names)]
	
	def writeWorkbook(self, index, datas, images):
		[self.ws.cell(row=index+2, column=x+1, value=data)for x, data in enumerate(datas)]
		self.ws.row_dimensions[index+2].height = Height*0.75
		for i, image in enumerate(images):
			column_leter = openpyxl.utils.cell.get_column_letter(self.ws.cell(row=index+2, column=(len(datas)+i+1)).column)
			self.ws.column_dimensions[column_leter].width = Width*0.13
			cell_address = self.ws.cell(row=index+2, column=len(datas)+1+i).coordinate
			img = openpyxl.drawing.image.Image(image)
			img.anchor = cell_address
			self.ws.add_image(img)
	
	def _saveWorkbook(self, path):
		self.wb.save(path)


if __name__ == '__main__':
	# grad cam の準備
	VGG = VGG()
	model = VGG.model
	model.load_state_dict(torch.load(modelPath, map_location='cuda:0'))
	gradcam = GradCam(model)
	excel = ExcelWriter()
	
	ImagePaths = list(chain.from_iterable([glob.glob('{}/*.jpg'.format(d)) for d in glob.glob('{}/*'.format(imgPaths))]))
	# ImagePaths = glob.glob(imgPaths+"/*")

	for i, path in enumerate(ImagePaths):
		gradImagePath = "resultImg\\"+str(i)+"_grad.jpg"
		# print(path)
		testImage = Image.open(path)
		testImageTensor = (transform((testImage))).unsqueeze(dim=0)
		imageSize = testImage.size
		featureImage = gradcam(testImageTensor).squeeze(dim=0)
		featureImage = transforms.ToPILImage()(featureImage)
		output = F.softmax(model(testImageTensor))
		# print("{:.6}".format(output.data[0, 0].item()))
		pred_idx = model(testImageTensor).max(1)[1]
		featureImage.save(gradImagePath)
		inputData = [os.path.basename(path), os.path.basename(os.path.dirname(path)), classList[pred_idx]] + [i.item() for i in output.data[0]]
		inputImage = [path, gradImagePath]
		excel.writeWorkbook(i, inputData, inputImage)
		# excel._saveWorkbook()
		# print(output.data.item())
	excel._saveWorkbook(resultFile)
